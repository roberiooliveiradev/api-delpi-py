# app/services/product_service.py
from app.repositories.product_repository import ProductRepository
from app.models.product_model import Product
from app.utils.logger import log_info, log_error
from app.core.exceptions import BusinessLogicError, DatabaseConnectionError
from typing import Optional

import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter


def get_product(code: str) -> Product:
    """
    Busca um produto no Protheus via repositório e retorna um modelo Pydantic.
    """
    repo = ProductRepository()
    log_info(f"Iniciando consulta do produto {code} via repositório")
    try:
        result = repo.get_product_by_code(code)
        return Product(**result)
    except BusinessLogicError as e:
        log_error(str(e))
        raise
    except Exception as e:
        log_error(f"Erro inesperado ao buscar produto {code}: {e}")
        raise DatabaseConnectionError(str(e))


def get_products(limit: int = 10) -> list[Product]:
    """
    Lista produtos do Protheus com limite definido.
    Retorna uma lista de modelos Product.
    """
    repo = ProductRepository()
    log_info(f"Listando até {limit} produtos do Protheus...")
    try:
        results = repo.list_products(limit)
        # Converte cada linha retornada em um objeto Product
        products = [Product(**r) for r in results]
        return products
    except Exception as e:
        log_error(f"Erro ao listar produtos: {e}")
        raise DatabaseConnectionError(str(e))

def search_products_by_description(
    description: str,
    page: int = 1,
    page_size: int = 50
) -> dict:
    repo = ProductRepository()
    log_info(f"Search only by description (page={page})")
    try:
        return repo.search_by_description(description, page, page_size)
    except Exception as e:
        log_error(f"Erro ao pesquisar produtos por descrição: {e}")
        raise DatabaseConnectionError(str(e))

def search_products(
    page: int = 1,
    page_size: int = 50,
    code: Optional[str] = None,
    description: Optional[str] = None,
    group: Optional[str] = None
) -> dict:
    repo = ProductRepository()
    log_info(f"Buscando produtos (page={page}) com filtros aplicados")
    try:
        return repo.search_products(page, page_size, code, description, group)
    except Exception as e:
        log_error(f"Erro ao pesquisar produtos: {e}")
        raise DatabaseConnectionError(str(e))


def get_structure(code: str, max_depth: int = 10, page: int = 1, page_size: int = 50) -> dict:
    repo = ProductRepository()
    log_info(f"Buscando estrutura (CTE) paginada para {code}")
    try:
        return repo.list_structure(code, max_depth, page, page_size)
    except Exception as e:
        log_error(f"Erro ao listar estrutura do produto {code}: {e}")
        raise DatabaseConnectionError(str(e))

def get_parents(code: str, max_depth: int = 10, page: int = 1, page_size: int = 50) -> dict:
    repo = ProductRepository()
    log_info(f"Buscando pais (CTE) paginados para {code}")
    try:
        return repo.list_parents(code, max_depth, page, page_size)
    except Exception as e:
        log_error(f"Erro ao listar produtos pai do item {code}: {e}")
        raise DatabaseConnectionError(str(e))

def get_suppliers(code: str, page: int = 1, page_size: int = 50) -> dict:
    repo = ProductRepository()
    log_info(f"Buscando fornecedores para {code} (página {page})")
    try:
        return repo.list_suppliers(code, page, page_size)
    except Exception as e:
        log_error(f"Erro ao listar fornecedores para {code}: {e}")
        raise DatabaseConnectionError(str(e))
    
def get_inbound_invoice_items(
    code: str,
    page: int = 1,
    page_size: int = 50,
    issue_date_start: Optional[str] = None,
    issue_date_end: Optional[str] = None,
    supplier: Optional[str] = None,
    branch: Optional[str] = None
) -> dict:
    repo = ProductRepository()
    log_info(f"Buscando NF-es de entrada de {code} (página {page})")
    try:
        return repo.list_inbound_invoice_items(code, page, page_size, issue_date_start, issue_date_end, supplier, branch)
    except Exception as e:
        log_error(f"Erro ao listar NF-es de entrada para {code}: {e}")
        raise DatabaseConnectionError(str(e))

def get_outbound_invoice_items(
    code: str,
    page: int = 1,
    page_size: int = 50,
    issue_date_start: Optional[str] = None,
    issue_date_end: Optional[str] = None,
    customer: Optional[str] = None,
    branch: Optional[str] = None
) -> dict:
    repo = ProductRepository()
    log_info(f"Buscando NF-es de saída de {code} (página {page})")
    try:
        return repo.list_outbound_invoice_items(code, page, page_size, issue_date_start, issue_date_end, customer, branch)
    except Exception as e:
        log_error(f"Erro ao listar NF-es de saída para {code}: {e}")
        raise DatabaseConnectionError(str(e))

def get_stock(
    code: str,
    page: int = 1,
    page_size: int = 50,
    branch: Optional[str] = None,
    location: Optional[str] = None
) -> dict:
    repo = ProductRepository()
    log_info(f"Buscando estoque para {code} (página {page})")

    try:
        return repo.list_stock(code, page, page_size, branch, location)
    except Exception as e:
        log_error(f"Erro ao listar estoque para {code}: {e}")
        raise DatabaseConnectionError(str(e))

def get_guide(
    code: str,
    page: int = 1,
    page_size: int = 50,
    branch: Optional[str] = None,
    max_depth: int = 10
) -> dict:
    repo = ProductRepository()
    log_info(f"Buscando estoque para {code} (página {page})")

    try:
        return repo.list_guide(code, page, page_size, branch, max_depth)
    except Exception as e:
        log_error(f"Erro ao listar estoque para {code}: {e}")
        raise DatabaseConnectionError(str(e))

def get_inspection(
    code: str,
    page: int = 1,
    page_size: int = 50,
    max_depth: int = 10
) -> dict:

    repo = ProductRepository()
    log_info(f"Buscando inspeções de processo para {code} (página {page})")

    try:
        return repo.list_inspection(code, page, page_size, max_depth)
    except Exception as e:
        log_error(f"Erro ao listar inspeções para {code}: {e}")
        raise DatabaseConnectionError(str(e))

def get_product_analyser(
    code: str,
    page: int = 1,
    page_size: int = 50,
    max_depth: int = 10
) -> dict:
    """
    Agrega todas as consultas de produto em uma única resposta:
    - dados gerais
    - estrutura (BOM)
    - roteiro (SG2)
    - inspeções (QP6/QP7/QP8)
    """
    repo = ProductRepository()
    log_info(f"Analisando produto completo {code}")

    try:
        product = repo.get_product_by_code(code)
        structure = repo.list_structure(code, max_depth, page, page_size)
        guide = repo.list_guide(code, page, page_size, None, max_depth)
        inspection = repo.list_inspection(code, page, page_size, max_depth)

        return {
            "success": True,
            "product": product,
            "structure": structure,
            "guide": guide,
            "inspection": inspection,
        }

    except Exception as e:
        log_error(f"Erro ao montar análise completa do produto {code}: {e}")
        raise DatabaseConnectionError(str(e))

def get_customers(code: str, page: int = 1, page_size: int = 50) -> dict:
    repo = ProductRepository()
    log_info(f"Buscando clientes amarrados ao produto {code} (página {page})")
    try:
        return repo.list_customers(code, page, page_size)
    except Exception as e:
        log_error(f"Erro ao listar clientes para o produto {code}: {e}")
        raise DatabaseConnectionError(str(e))


# --------------------------------------------------------------------
# ESTRUTURA EM TABELA DE EXCEL
# --------------------------------------------------------------------
def get_structure_excel2(code: str) -> io.BytesIO:
    """
    Gera a planilha Excel no formato oficial DELPI:
    - Produto acabado no topo
    - Intermediários agrupados
    - Componentes formatados conforme norma
    - Fonte vermelha somente para matérias-primas (MP) com unidade PC e QTD > 2 (em milheiro)
    """
    repo = ProductRepository()
    log_info(f"Gerando planilha Excel hierárquica e formatada para {code}")

    # structure = repo.list_structure(code, max_depth, 1, 500)
    structure = repo.list_structure_full(code)
    root = structure["data"]

    rows = []
    meta_map = {}  # Dicionário auxiliar: {componente_code: {"type": ..., "unit": ...}}

    # ---------------------------------------------------------------------
    # Cabeçalho do produto pai (sempre aparece como 1ª linha)
    # ---------------------------------------------------------------------
    # 🔧 Só cria a linha do PA raiz se NÃO houver MP direta
    has_direct_mp = any(c.get("type") == "MP" for c in root.get("components", []))

    if not has_direct_mp:
        rows.append([
            root.get("code", ""),
            root.get("description", ""),
            "",
            "",
            "",
            "",
            root.get("type", ""),
            root.get("unit", "")
        ])


    # ---------------------------------------------------------------------
    # Função recursiva:
    # - PA/PI funcionam apenas como "pai" (Código / Descrição)
    # - SOMENTE MPs ocupam Item, QTD, Componente, Descrição (componente)
    # ---------------------------------------------------------------------
    def process_components(parent_code: str, parent_desc: str, components: list[dict]):
        for comp in components:
            code_comp = comp.get("code", "")
            desc_comp = comp.get("description", "")
            comp_type = comp.get("type", "")
            comp_unit = comp.get("unit", "")

            if comp_type == "MP":
                # Linha de saída: pai = PA/PI imediato (ou o próprio PA raiz se MP direta)
                comp_item = comp.get("item", "")
                comp_qtd = comp.get("quantity", "")

                rows.append([
                    parent_code,      # Código do PA/PI agrupador
                    parent_desc,      # Descrição do PA/PI agrupador
                    comp_item,        # Item (apenas MP)
                    comp_qtd,         # QTD (apenas MP, será tratada depois)
                    code_comp,        # Código da MP
                    desc_comp,        # Descrição da MP
                    comp_type,
                    comp_unit,
                ])
                meta_map[code_comp] = {"type": comp_type, "unit": comp_unit}

            else:
                # PA / PI → desce um nível e usa este como novo "pai"
                if comp.get("components"):
                    process_components(code_comp, desc_comp, comp["components"])

    if root.get("components"):
        process_components(root.get("code", ""), root.get("description", ""), root["components"])

    # ---------------------------------------------------------------------
    # Geração da coluna "Item" (A, B, C... fixo por componente)
    # ---------------------------------------------------------------------
    from string import ascii_uppercase

    def generate_item_label(index):
        """Retorna rótulo tipo Excel: A-Z, AA, AB, etc."""
        letters = ""
        while True:
            index, remainder = divmod(index, 26)
            letters = ascii_uppercase[remainder] + letters
            if index == 0:
                break
            index -= 1
        return letters

    # Dicionário que garante o mesmo item para o mesmo componente
    item_map = {}
    item_counter = 0

    for r in rows:
        comp_code = r[4]
        if comp_code:
            if comp_code not in item_map:
                item_map[comp_code] = generate_item_label(item_counter)
                item_counter += 1
            # Atribui o item à posição 2 (coluna Item) somente para linhas de MP
            r[2] = item_map[comp_code]

    # ---------------------------------------------------------------------
    # Criação da planilha Excel
    # ---------------------------------------------------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "Estrutura DELPI"

    headers = ["Código", "Descrição", "Item", "QTD", "Componente", "Descrição"]
    ws.append(headers)
    fontName = "Arial Narrow"
    fontSize = 10
    blackColor = "000000"
    blueColor = "0000FF"
    redColor = "FF0000"

    thin = Side(border_style="thin", color=blackColor)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    header_font = Font(bold=True, color=blueColor, name=fontName, size=fontSize)
    normal_font = Font(color=blackColor, name=fontName, size=fontSize)
    red_font = Font(color=redColor, name=fontName, size=fontSize)
    blue_font = Font(color=blueColor, name=fontName, size=fontSize)

    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.alignment = align_center
        c.border = border
        ws.column_dimensions[get_column_letter(col)].width = 50 if col in [2, 6] else 14

    for r in rows:
        ws.append(r[:6])  # mantém apenas as colunas exibidas

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for c in row:
            c.border = border
            c.alignment = align_left if c.column in [2, 6] else align_center
            c.font = normal_font

    # ---------------------------------------------------------------------
    # Mesclar colunas "Código" e "Descrição" por agrupamento
    # ---------------------------------------------------------------------
    current_parent = None
    start_row = 2
    for i in range(2, ws.max_row + 1):
        parent = ws.cell(row=i, column=1).value
        if parent != current_parent:
            if current_parent is not None and start_row < i - 1:
                ws.merge_cells(start_row=start_row, start_column=1, end_row=i - 1, end_column=1)
                ws.merge_cells(start_row=start_row, start_column=2, end_row=i - 1, end_column=2)
            start_row = i
            current_parent = parent
    if current_parent and start_row < ws.max_row:
        ws.merge_cells(start_row=start_row, start_column=1, end_row=ws.max_row, end_column=1)
        ws.merge_cells(start_row=start_row, start_column=2, end_row=ws.max_row, end_column=2)

    # ---------------------------------------------------------------------
    # Aplicar ajustes de quantidade e formatação de fonte
    # (somente linhas de MP, pois só elas têm componente e meta_map)
    # ---------------------------------------------------------------------
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        qtd_cell = row[3].value
        comp_code = str(row[4].value or "").strip()

        comp_meta = meta_map.get(comp_code, {})
        comp_type = comp_meta.get("type", "")
        comp_unit = comp_meta.get("unit", "")

        try:
            qtd = float(str(qtd_cell).replace(",", ".")) if qtd_cell not in [None, ""] else 0
        except Exception:
            qtd = 0

        # ❗ SE NÃO É MP → NÃO ALTERA QTD, NÃO PINTA, NÃO FORMATA
        if not (comp_code == "" or comp_code not in meta_map):
            # 🔧 Ajuste de quantidade conforme unidade (somente MP)
            if comp_unit == "PC":
                if qtd % 1000 != 0:
                    qtd = 1
                else:
                    qtd = qtd / 1000
            else:
                qtd = 1

            row[3].value = qtd

        # 🔵 Cor base azul
        for c in row[0:6]:
            c.font = blue_font

        # 🔴 Regras de destaque: MP + PC + QTD >= 2
        if comp_unit == "PC" and qtd >= 2:
            for c in row[2:6]:
                c.font = red_font


    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


# --------------------------------------------------------------------
# ESTRUTURA EM TABELA DE EXCEL (COM PARENT_FACTOR ACUMULATIVO)
# --------------------------------------------------------------------
def get_structure_excel(code: str) -> io.BytesIO:
    """
    Gera a planilha Excel no formato oficial DELPI:
    - Produto acabado no topo
    - Intermediários agrupados
    - Matérias-primas com quantidade final acumulada
    - parent_factor acumulativo: PA(1) → PI → MP
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from string import ascii_uppercase

    repo = ProductRepository()
    log_info(f"Gerando planilha Excel hierárquica e formatada para {code}")

    structure = repo.list_structure_full(code)
    root = structure["data"]

    rows: list[list] = []
    meta_map: dict[str, dict] = {}

    # ------------------------------------------------------------------
    # Cabeçalho do produto pai (somente se NÃO houver MP direta)
    # ------------------------------------------------------------------
    has_direct_mp = any(c.get("type") == "MP" for c in root.get("components", []))

    if not has_direct_mp:
        rows.append([
            root.get("code", ""),
            root.get("description", ""),
            "",
            "",
            "",
            "",
            root.get("type", ""),
            root.get("unit", ""),
            1.0,  # parent_factor do PA raiz
        ])

    # ------------------------------------------------------------------
    # Função recursiva com parent_factor acumulativo
    # ------------------------------------------------------------------
    def process_components(
        parent_code: str,
        parent_desc: str,
        components: list[dict],
        parent_factor: float = 1.0
    ):
        for comp in components:
            code_comp = comp.get("code", "")
            desc_comp = comp.get("description", "")
            comp_type = comp.get("type", "")
            comp_unit = comp.get("unit", "")
            comp_qtd = float(comp.get("quantity", 1) or 1)

            if comp_type == "MP":
                rows.append([
                    parent_code,
                    parent_desc,
                    comp.get("item", ""),
                    comp_qtd,          # qtd original
                    code_comp,
                    desc_comp,
                    comp_type,
                    comp_unit,
                    parent_factor,     # 🔑 fator acumulado
                ])
                meta_map[code_comp] = {"type": comp_type, "unit": comp_unit}

            else:
                # PA / PI → acumula fator e desce
                new_parent_factor = parent_factor * comp_qtd
                if comp.get("components"):
                    process_components(
                        code_comp,
                        desc_comp,
                        comp["components"],
                        new_parent_factor
                    )

    if root.get("components"):
        process_components(
            root.get("code", ""),
            root.get("description", ""),
            root["components"],
            parent_factor=1.0
        )

    # ------------------------------------------------------------------
    # Geração da coluna Item (A, B, C...)
    # ------------------------------------------------------------------
    def generate_item_label(index: int) -> str:
        letters = ""
        while True:
            index, remainder = divmod(index, 26)
            letters = ascii_uppercase[remainder] + letters
            if index == 0:
                break
            index -= 1
        return letters

    item_map: dict[str, str] = {}
    item_counter = 0

    for r in rows:
        comp_code = r[4]
        if comp_code:
            if comp_code not in item_map:
                item_map[comp_code] = generate_item_label(item_counter)
                item_counter += 1
            r[2] = item_map[comp_code]

    # ------------------------------------------------------------------
    # Criação da planilha Excel
    # ------------------------------------------------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "Estrutura DELPI"

    headers = ["Código", "Descrição", "Item", "QTD", "Componente", "Descrição"]
    ws.append(headers)

    font_name = "Arial Narrow"
    font_size = 10

    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    header_font = Font(bold=True, color="0000FF", name=font_name, size=font_size)
    normal_font = Font(color="000000", name=font_name, size=font_size)
    red_font = Font(color="FF0000", name=font_name, size=font_size)
    blue_font = Font(color="0000FF", name=font_name, size=font_size)

    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.alignment = align_center
        c.border = border
        ws.column_dimensions[get_column_letter(col)].width = 50 if col in [2, 6] else 14

    for r in rows:
        ws.append(r[:6])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for c in row:
            c.border = border
            c.alignment = align_left if c.column in [2, 6] else align_center
            c.font = normal_font

    # ------------------------------------------------------------------
    # Mesclar Código / Descrição por agrupamento
    # ------------------------------------------------------------------
    current_parent = None
    start_row = 2

    for i in range(2, ws.max_row + 1):
        parent = ws.cell(row=i, column=1).value
        if parent != current_parent:
            if current_parent is not None and start_row < i - 1:
                ws.merge_cells(start_row=start_row, start_column=1, end_row=i - 1, end_column=1)
                ws.merge_cells(start_row=start_row, start_column=2, end_row=i - 1, end_column=2)
            start_row = i
            current_parent = parent

    if current_parent and start_row < ws.max_row:
        ws.merge_cells(start_row=start_row, start_column=1, end_row=ws.max_row, end_column=1)
        ws.merge_cells(start_row=start_row, start_column=2, end_row=ws.max_row, end_column=2)

    # ------------------------------------------------------------------
    # Ajuste FINAL da quantidade (MP + parent_factor)
    # ------------------------------------------------------------------
    for idx, row in enumerate(
        ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)),
        start=0
    ):
        qtd_cell = row[3].value
        comp_code = str(row[4].value or "").strip()

        comp_meta = meta_map.get(comp_code)
        if not comp_meta:
            for c in row:
                c.font = blue_font
            continue

        comp_unit = comp_meta["unit"]

        try:
            qtd = float(str(qtd_cell).replace(",", "."))
        except Exception:
            qtd = 0

        parent_factor = rows[idx][8] if len(rows[idx]) > 8 else 1

        if comp_unit == "PC":
            qtd = qtd / 1000 if qtd % 1000 == 0 else 1
        else:
            qtd = 1

        qtd_final = qtd * parent_factor
        row[3].value = qtd_final

        # 🔵 Cor base azul
        for c in row:
            c.font = blue_font

        # 🔴 Destaque: MP + PC + QTD >= 2
        if qtd_final >= 2:
            for c in row[2:6]:
                c.font = red_font

    # ------------------------------------------------------------------
    # Retorno do Excel
    # ------------------------------------------------------------------
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
